"""
DVends Replenishment Report Exporter
Usage: python open_vending.py [--headless]
"""

import asyncio
import datetime
import hashlib
import json
import os
import sqlite3
import sys
from pathlib import Path
from playwright.async_api import async_playwright

USERNAME = os.environ.get('OV_USERNAME', '')
PASSWORD = os.environ.get('OV_PASSWORD', '')

DB_DIR  = Path(__file__).parent / "db"
LOG_DIR = Path(__file__).parent / "log"
DB_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
_log_file = open(LOG_DIR / "scan.log", "a", encoding="utf-8")

HEADLESS   = "--headless" in sys.argv
LOGIN_ONLY = "--login-only" in sys.argv

BROWSER_STOP = DB_DIR / ".browser_stop"

_DEFAULT_LOGIN_URL = "https://vendingportal.azurewebsites.net/SuperAdmin/SPLogin.aspx"
LOGIN_URL  = os.environ.get('OV_LANDING_URL') or _DEFAULT_LOGIN_URL
REPORT_URL = "https://vendingportal.azurewebsites.net/SuperAdmin/SPReplenishmentV2.aspx"


SQLITE_DB = DB_DIR / "vending.db"


def status(msg):
    print(f"STATUS: {msg}", flush=True)
    print(f"{datetime.datetime.now():%Y-%m-%d %H:%M:%S} {msg}", file=_log_file, flush=True)


def import_to_sqlite(xlsx_path):
    import shutil
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    now = datetime.datetime.now().isoformat()

    conn = sqlite3.connect(str(SQLITE_DB))
    conn.execute("PRAGMA journal_mode=WAL")
    SCHEMA_VERSION = 4
    if conn.execute("PRAGMA user_version").fetchone()[0] < SCHEMA_VERSION:
        # Bootstrap for a fresh/empty vending.db only. The normal path is
        # migrate_to_vending.py, which creates these tables and sets
        # user_version=4; this block must never drop existing data.
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS current_state (
                machine      TEXT NOT NULL,
                lane         TEXT NOT NULL,
                restock      INTEGER,
                updated_at   TEXT,
                pid          TEXT,
                product_name TEXT,
                bal          INTEGER,
                lane_size    INTEGER,
                PRIMARY KEY (machine, lane)
            );
            CREATE TABLE IF NOT EXISTS change_log (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                detected_at  TEXT NOT NULL,
                machine      TEXT NOT NULL,
                lane         TEXT NOT NULL,
                old_restock  INTEGER,
                new_restock  INTEGER
            );
            CREATE TABLE IF NOT EXISTS lane_events (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                detected_at  TEXT NOT NULL,
                machine      TEXT NOT NULL,
                lane         TEXT NOT NULL,
                pid          TEXT,
                product_name TEXT,
                old_bal      INTEGER,
                new_bal      INTEGER,
                old_restock  INTEGER,
                new_restock  INTEGER,
                lane_size    INTEGER
            );
            CREATE INDEX IF NOT EXISTS idx_lane_events_mpd ON lane_events (machine, pid, detected_at);
        """)
        conn.execute(f"PRAGMA user_version = {SCHEMA_VERSION}")
        conn.commit()

    def to_int(v):
        if v is None:
            return None
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    changes = []
    for ws in wb.worksheets:
        machine = ws.title
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # locate columns from header row (report: No. | Product ID |
        # Product Name | Bal Qty | Lane Size | Restock)
        header = [str(h).strip().lower() if h else '' for h in rows[0]]
        def col(name):
            try:
                return header.index(name)
            except ValueError:
                return None
        restock_idx = col('restock')
        if restock_idx is None:
            continue
        pid_idx  = col('product id')
        name_idx = col('product name')
        bal_idx  = col('bal qty')
        size_idx = col('lane size')

        def cell(row, idx):
            return row[idx] if idx is not None and idx < len(row) else None

        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            lane = str(row[0])
            restock = to_int(cell(row, restock_idx))
            if restock is None:
                continue
            pid       = cell(row, pid_idx)
            pid       = str(int(pid)) if isinstance(pid, (int, float)) else (str(pid).strip() if pid else None)
            pname     = str(cell(row, name_idx)).strip() if cell(row, name_idx) else None
            bal       = to_int(cell(row, bal_idx))
            lane_size = to_int(cell(row, size_idx))

            existing = conn.execute(
                "SELECT restock, pid, bal FROM current_state WHERE machine=? AND lane=?",
                (machine, lane)
            ).fetchone()

            if existing is None:
                conn.execute(
                    "INSERT INTO current_state (machine, lane, restock, updated_at, pid, product_name, bal, lane_size) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    (machine, lane, restock, now, pid, pname, bal, lane_size)
                )
                continue

            old_restock, old_pid, old_bal = existing
            restock_changed = old_restock != restock
            # pid/bal migrated rows start as NULL — filling them in is a
            # backfill, not a real event
            pid_changed = old_pid is not None and pid is not None and old_pid != pid
            bal_changed = old_bal is not None and bal is not None and old_bal != bal

            if restock_changed:
                conn.execute(
                    "INSERT INTO change_log (detected_at, machine, lane, old_restock, new_restock) VALUES (?,?,?,?,?)",
                    (now, machine, lane, old_restock, restock)
                )
                changes.append({'machine': machine, 'lane': lane, 'old': old_restock, 'new': restock})

            if restock_changed or pid_changed or bal_changed:
                conn.execute(
                    "INSERT INTO lane_events (detected_at, machine, lane, pid, product_name, "
                    "old_bal, new_bal, old_restock, new_restock, lane_size) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (now, machine, lane, pid, pname, old_bal, bal, old_restock, restock, lane_size)
                )
                conn.execute(
                    "UPDATE current_state SET restock=?, updated_at=?, pid=?, product_name=?, bal=?, lane_size=? "
                    "WHERE machine=? AND lane=?",
                    (restock, now, pid, pname, bal, lane_size, machine, lane)
                )
            elif old_pid is None or old_bal is None:
                # silent backfill of newly captured columns (no updated_at bump)
                conn.execute(
                    "UPDATE current_state SET pid=?, product_name=?, bal=?, lane_size=? WHERE machine=? AND lane=?",
                    (pid, pname, bal, lane_size, machine, lane)
                )

    wb.close()

    # refresh realtime estimates in the derived daily_sales layer
    try:
        sys.path.insert(0, str(Path(__file__).parent / "src"))
        import daily_sales
        daily_sales.refresh_est(conn)
    except Exception as e:
        status(f"daily_sales refresh failed: {e}")

    conn.commit()
    conn.close()

    # keep a persistent copy for next startup, then archive
    shutil.copy2(str(xlsx_path), str(DB_DIR / "last_report.xlsx"))
    tmp_dir = DB_DIR / "tmp"
    tmp_dir.mkdir(exist_ok=True)
    for f in DB_DIR.glob("*.xlsx"):
        if f.name != "last_report.xlsx":
            shutil.move(str(f), str(tmp_dir / f.name))

    # prune archived reports older than 7 days — their data already lives
    # in change_log / lane_events / daily_sales
    import time
    cutoff = time.time() - 7 * 86400
    for f in tmp_dir.glob("*.xlsx"):
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
        except OSError:
            pass

    status(f"DB: {len(changes)} restock change(s) detected")
    return changes


async def watch_f9(page, stop_event):
    trigger = DB_DIR / ".f9_trigger"
    LOG_DIR.mkdir(exist_ok=True)
    while not stop_event.is_set():
        if trigger.exists():
            try:
                trigger.unlink()
                html = await page.content()
                out = LOG_DIR / "dom_debug.html"
                out.write_text(html, encoding='utf-8')
                status(f"F9: DOM saved -> {out}")
            except Exception as e:
                status(f"F9: capture failed: {e}")
        await asyncio.sleep(0.3)


async def login(page):
    status("Navigating to login page...")
    await page.goto(LOGIN_URL)
    await page.wait_for_load_state("domcontentloaded")

    for sel in ["#txtUsername", "input[name='txtUsername']", "input[type='text']"]:
        loc = page.locator(sel).first
        if await loc.count() > 0:
            await loc.fill(USERNAME)
            break

    for sel in ["#txtPassword", "input[name='txtPassword']", "input[type='password']"]:
        loc = page.locator(sel).first
        if await loc.count() > 0:
            await loc.fill(PASSWORD)
            break

    for sel in ["#btnLogin", "input[value='Login']", "input[type='submit']", "button[type='submit']"]:
        loc = page.locator(sel).first
        if await loc.count() > 0:
            await loc.click()
            break

    await page.wait_for_url("**/SPDashboard.aspx", timeout=15000)
    status("Login successful")


async def export_excel(page):
    status("Loading replenishment report...")
    await page.goto(REPORT_URL)
    await page.wait_for_load_state("networkidle")

    status("Exporting Excel...")
    btn = None
    for sel in [
        "input[value='Export Excel']", "#Exporttoexcel",
        "input[value*='Excel']",
    ]:
        loc = page.locator(sel).first
        if await loc.count() > 0:
            btn = loc
            status(f"Export button found: {sel}")
            break
    if btn is None:
        (LOG_DIR / "export_fail.html").write_text(await page.content(), encoding="utf-8")
        await page.screenshot(path=str(LOG_DIR / "export_fail.png"), full_page=True)
        raise RuntimeError("Export button not found — page saved to log/export_fail.html/.png")

    try:
        async with page.expect_download(timeout=120000) as dl:
            await btn.click()
    except Exception:
        try:
            import re
            html = await page.content()
            (LOG_DIR / "export_fail.html").write_text(html, encoding="utf-8")
            await page.screenshot(path=str(LOG_DIR / "export_fail.png"), full_page=True)
            m = re.search(r"<title>(.*?)</title>", html, re.S)
            if m and m.group(1).strip():
                status(f"Portal server error: {m.group(1).strip()}")
            status("Export failed — page saved to log/export_fail.html/.png")
        except Exception:
            pass
        raise

    download = await dl.value
    filename = download.suggested_filename or "replenishment_report.xlsx"
    save_path = DB_DIR / filename
    await download.save_as(str(save_path))

    status(f"Saved to {save_path}")
    return save_path


async def main():
    xlsx_path = None
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        stop_event = asyncio.Event()
        f9_task = None
        if not HEADLESS:
            f9_task = asyncio.create_task(watch_f9(page, stop_event))

        await login(page)

        if LOGIN_ONLY:
            status("Browser ready — waiting")
            while not BROWSER_STOP.exists():
                await asyncio.sleep(0.5)
            try: BROWSER_STOP.unlink()
            except: pass
        else:
            xlsx_path = await export_excel(page)

        stop_event.set()
        if f9_task:
            await f9_task

        await browser.close()

        if not HEADLESS and not LOGIN_ONLY:
            input("\nPress Enter to exit...")

    if xlsx_path:
        diffs = import_to_sqlite(xlsx_path)
        print(f"FILE: {DB_DIR / 'last_report.xlsx'}", flush=True)
        if diffs:
            print(f"DIFFS: {json.dumps(diffs, ensure_ascii=False)}", flush=True)


try:
    asyncio.run(main())
except Exception:
    import traceback
    tb = traceback.format_exc()
    print(f"{datetime.datetime.now():%Y-%m-%d %H:%M:%S} CRASH:\n{tb}", file=_log_file, flush=True)
    sys.stderr.write(tb)
    sys.exit(1)
